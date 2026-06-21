"""
policy_brief.py — Ministry-ready Excel workbook export for the Speed Safety Score pipeline.

Exports all Critical and High Risk segments across 8 sheets:
  1. Executive Summary
  2. Critical Segments
  3. High Risk Segments
  4. Summary by Road Class
  5. Thailand — Province Risk Summary
  6. Maharashtra — District Risk Summary  (spatial join with OSM admin boundaries)
  7. Intervention Zones (corridors)
  8. Methodology Note
"""

import numpy as np
import pandas as pd
import geopandas as gpd
from pathlib import Path

THAILAND_PROVINCES = {
    10: "Bangkok", 11: "Samut Prakan", 12: "Nonthaburi", 13: "Pathum Thani",
    14: "Phra Nakhon Si Ayutthaya", 15: "Ang Thong", 16: "Lop Buri",
    17: "Sing Buri", 18: "Chai Nat", 19: "Saraburi", 20: "Chon Buri",
    21: "Rayong", 22: "Chanthaburi", 23: "Trat", 24: "Chachoengsao",
    25: "Prachin Buri", 26: "Nakhon Nayok", 27: "Sa Kaeo",
    30: "Nakhon Ratchasima", 31: "Buri Ram", 32: "Surin", 33: "Si Sa Ket",
    34: "Ubon Ratchathani", 35: "Yasothon", 36: "Chaiyaphum",
    37: "Amnat Charoen", 38: "Bueng Kan", 39: "Nong Bua Lam Phu",
    40: "Khon Kaen", 41: "Udon Thani", 42: "Loei", 43: "Nong Khai",
    44: "Maha Sarakham", 45: "Roi Et", 46: "Kalasin", 47: "Sakon Nakhon",
    48: "Nakhon Phanom", 49: "Mukdahan", 50: "Chiang Mai", 51: "Lamphun",
    52: "Lampang", 53: "Uttaradit", 54: "Phrae", 55: "Nan", 56: "Phayao",
    57: "Chiang Rai", 58: "Mae Hong Son", 60: "Nakhon Sawan",
    61: "Uthai Thani", 62: "Kamphaeng Phet", 63: "Tak", 64: "Sukhothai",
    65: "Phitsanulok", 66: "Phichit", 67: "Phetchabun", 70: "Ratchaburi",
    71: "Kanchanaburi", 72: "Suphan Buri", 73: "Nakhon Pathom",
    74: "Samut Sakhon", 75: "Samut Songkhram", 76: "Phetchaburi",
    77: "Prachuap Khiri Khan", 80: "Nakhon Si Thammarat", 81: "Krabi",
    82: "Phangnga", 83: "Phuket", 84: "Surat Thani", 85: "Ranong",
    86: "Chumphon", 90: "Songkhla", 91: "Satun", 92: "Trang",
    93: "Phatthalung", 94: "Pattani", 95: "Yala", 96: "Narathiwat",
}


def _infer_jurisdiction(road_class: str, country: str) -> str:
    """Infer the responsible road authority from road class and country."""
    rc = str(road_class).lower()
    if "maharashtra" in country.lower() or country.upper() in ("MH", "IN"):
        return {
            "motorway":    "NHAI (National Highways Authority of India)",
            "trunk":       "NHAI / MSRDC (Maharashtra State Road Dev. Corp.)",
            "primary":     "Maharashtra PWD (Public Works Dept.)",
            "secondary":   "Maharashtra PWD / District Authority",
            "tertiary":    "District / Municipal Authority",
            "residential": "Municipal Corporation / Local Authority",
        }.get(rc, "Maharashtra PWD")
    else:  # Thailand
        return {
            "motorway":    "DOH — Department of Highways",
            "trunk":       "DOH — Department of Highways",
            "primary":     "DOH — Department of Highways",
            "secondary":   "DRR — Department of Rural Roads",
            "tertiary":    "DRR / LAO — Local Administration",
            "residential": "LAO — Local Administration Organization",
        }.get(rc, "DOH — Department of Highways")


def _build_segment_row(r, rank: int) -> dict:
    """Build one policy brief row from a GeoDataFrame row."""
    speed_limit   = r.get("speed_limit", np.nan)
    rec_limit     = r.get("recommended_limit", r.get("ss_limit", np.nan))
    sss           = r.get("sss", np.nan)
    speed_85th    = r.get("speed_85th", np.nan)
    pct_over      = r.get("pct_over_limit", np.nan)
    sinuosity     = r.get("sinuosity", np.nan)
    ntl_score     = r.get("ntl_exposure_score", np.nan)
    sss_band      = r.get("sss_band", "")
    priority_band = r.get("priority_band", sss_band)
    nilsson       = r.get("nilsson_fatal_ratio", np.nan)
    credibility   = r.get("credibility_class", "")
    osm_lit       = str(r.get("osm_lit", "") or "")
    osm_surface   = str(r.get("osm_surface", "") or "")
    blindspot     = bool(r.get("mapillary_blindspot", False))
    land_use      = str(r.get("land_use", "") or "")
    road_class    = str(r.get("road_class_norm", r.get("road_class", "")) or "")
    country       = str(r.get("country", r.get("country_code", "")) or "")

    # Intervention actions
    actions = []
    if pd.notna(speed_limit) and pd.notna(rec_limit) and speed_limit - rec_limit > 20:
        actions.append(f"Reduce speed limit to {rec_limit:.0f} km/h (major revision + enforcement)")
    elif pd.notna(speed_limit) and pd.notna(rec_limit) and speed_limit > rec_limit:
        actions.append(f"Reduce speed limit to {rec_limit:.0f} km/h")
    if pd.notna(sinuosity) and sinuosity >= 1.5:
        actions.append("Install curve warning chevrons and advance warning signs")
    elif pd.notna(sinuosity) and sinuosity >= 1.2:
        actions.append("Install curve advisory speed signs")
    if blindspot:
        actions.append("Deploy speed camera or automated enforcement (unmonitored segment)")
    if osm_lit == "no" and land_use in ("urban", "interurban"):
        actions.append("Install street lighting (confirmed unlit road in populated area)")
    elif pd.notna(ntl_score) and ntl_score > 60 and osm_lit not in ("yes",):
        actions.append("Assess street lighting — high nighttime pedestrian activity detected via VIIRS")
    if osm_surface in ("unpaved", "gravel", "dirt", "compacted", "ground"):
        actions.append("Resurface to sealed asphalt (unpaved surface — loss-of-control risk)")
    if pd.notna(nilsson) and nilsson > 4 and road_class in ("trunk", "primary"):
        actions.append("Install median barrier / physical separation")
    if credibility == "Non-Credible":
        actions.append("Redesign limit scheme — posted limit widely ignored by drivers")
    if road_class == "residential" and pd.notna(speed_limit) and speed_limit > 30:
        actions.append("Implement traffic calming (residential road)")
    if not actions:
        actions.append("Monitor and schedule audit")

    # Why dangerous
    reasons = []
    if pd.notna(sss) and sss >= 50:
        reasons.append(f"SSS {sss:.0f}/100")
    if pd.notna(speed_85th) and pd.notna(speed_limit) and speed_85th > speed_limit + 10:
        reasons.append(f"85th-pct {speed_85th:.0f} > limit {speed_limit:.0f} km/h")
    if pd.notna(pct_over) and pct_over > 40:
        reasons.append(f"{pct_over:.0f}% exceed limit")
    if pd.notna(nilsson) and nilsson > 2:
        reasons.append(f"{nilsson:.1f}x fatal crash risk (Nilsson)")
    if pd.notna(sinuosity) and sinuosity >= 1.20:
        reasons.append(f"Sinuous geometry (SI={sinuosity:.2f})")
    if pd.notna(ntl_score) and ntl_score > 60:
        reasons.append(f"High nighttime activity (NTL={ntl_score:.0f})")
    if credibility == "Non-Credible":
        reasons.append("Limit non-credible — drivers ignore it")

    return {
        "Priority Rank":             rank,
        "Segment ID":                r.get("segment_id", ""),
        "Road Name":                 r.get("road_name", r.get("english_road", "")),
        "Country":                   country,
        "Road Class":                road_class,
        "Responsible Authority":     _infer_jurisdiction(road_class, country),
        "Land Use":                  land_use,
        "Posted Limit (km/h)":       speed_limit,
        "Recommended Limit (km/h)":  rec_limit,
        "Reduction Needed (km/h)":   (speed_limit - rec_limit
                                      if pd.notna(speed_limit) and pd.notna(rec_limit) else np.nan),
        "Speed Safety Score":        round(sss, 1) if pd.notna(sss) else "",
        "SSS Band":                  sss_band,
        "Priority Band":             priority_band,
        "Why Dangerous":             "; ".join(reasons) if reasons else sss_band,
        "Intervention Actions":      "; ".join(actions),
        "85th Pct Speed (km/h)":     round(speed_85th, 1) if pd.notna(speed_85th) else "",
        "% Vehicles Over Limit":     round(pct_over, 1) if pd.notna(pct_over) else "",
        "Nilsson Fatal Risk Ratio":  round(nilsson, 2) if pd.notna(nilsson) else "",
        "Sinuosity Index":           round(sinuosity, 3) if pd.notna(sinuosity) else "",
        "VIIRS NTL Score (0-100)":   round(ntl_score, 1) if pd.notna(ntl_score) else "",
        "OSM Surface":               osm_surface if osm_surface else "unknown",
        "OSM Lit":                   osm_lit if osm_lit else "unknown",
        "Mapillary Blindspot":       "Yes — deploy enforcement" if blindspot else "No",
        "Credibility":               credibility,
        "Street View Link":          r.get("image_url", ""),
    }


def _build_admin_summary(gdf: gpd.GeoDataFrame) -> tuple:
    """
    Returns (th_df, mh_df):
      th_df — Thailand segments aggregated by province (province_id lookup)
      mh_df — Maharashtra segments spatially joined to OSM district boundaries
    Both DataFrames contain: admin unit, segment count, avg SSS, Critical/High Risk counts,
    top road class, responsible authority hint.
    """
    band_col = "sss_band" if "sss_band" in gdf.columns else "priority_band"
    country_col = "country" if "country" in gdf.columns else "country_code"

    def _summarise(grp_df, name_col, authority_hint):
        rows = []
        for unit, g in grp_df.groupby(name_col):
            n      = len(g)
            avg_ss = g["sss"].mean() if "sss" in g.columns else np.nan
            n_crit = (g[band_col] == "Critical").sum()   if band_col in g.columns else 0
            n_high = (g[band_col] == "High Risk").sum()  if band_col in g.columns else 0
            n_prio = n_crit + n_high
            rc_col = "road_class_norm" if "road_class_norm" in g.columns else "road_class"
            top_rc = g[rc_col].mode()[0] if rc_col in g.columns and len(g) else ""
            rows.append({
                name_col:                    unit,
                "Total Segments":            n,
                "Critical":                  int(n_crit),
                "High Risk":                 int(n_high),
                "Priority Segments":         int(n_prio),
                "% Priority":                f"{100*n_prio/n:.1f}%" if n else "",
                "Avg SSS":                   round(avg_ss, 1) if pd.notna(avg_ss) else "",
                "Most Common Road Class":    top_rc,
                "Responsible Authority":     authority_hint,
            })
        return pd.DataFrame(rows).sort_values("Avg SSS", ascending=False).reset_index(drop=True)

    # ── Thailand: group by province_id ───────────────────────────────────────
    th_df = pd.DataFrame()
    th_mask = gdf[country_col].str.contains("Thailand", case=False, na=False)
    th = gdf[th_mask].copy()
    if len(th) and "province_id" in th.columns:
        th["province_id_n"] = pd.to_numeric(th["province_id"], errors="coerce")
        th["Province"] = th["province_id_n"].map(
            lambda x: THAILAND_PROVINCES.get(int(x), f"Province {int(x)}")
            if pd.notna(x) else "Unknown"
        )
        th_df = _summarise(th, "Province", "DOH / DRR — Department of Highways or Rural Roads")
        th_df.insert(1, "Province ID", th.groupby("Province")["province_id_n"].first().reindex(th_df["Province"]).values)
        th_df.insert(2, "Region", th_df["Province"].map({
            **{p: "Central"  for p in ["Bangkok","Nonthaburi","Pathum Thani","Samut Prakan","Samut Sakhon","Samut Songkhram","Nakhon Pathom","Suphan Buri","Ratchaburi","Kanchanaburi","Phetchaburi","Prachuap Khiri Khan","Chai Nat","Lop Buri","Sing Buri","Ang Thong","Phra Nakhon Si Ayutthaya","Saraburi","Nakhon Nayok","Chachoengsao","Prachin Buri","Sa Kaeo","Uthai Thani","Kamphaeng Phet"]},
            **{p: "North"    for p in ["Chiang Mai","Chiang Rai","Lamphun","Lampang","Phrae","Nan","Phayao","Mae Hong Son","Uttaradit","Tak","Sukhothai","Phitsanulok","Phichit","Phetchabun","Nakhon Sawan"]},
            **{p: "Northeast" for p in ["Nakhon Ratchasima","Buri Ram","Surin","Si Sa Ket","Ubon Ratchathani","Yasothon","Chaiyaphum","Amnat Charoen","Bueng Kan","Nong Bua Lam Phu","Khon Kaen","Udon Thani","Loei","Nong Khai","Maha Sarakham","Roi Et","Kalasin","Sakon Nakhon","Nakhon Phanom","Mukdahan"]},
            **{p: "East"     for p in ["Chon Buri","Rayong","Chanthaburi","Trat"]},
            **{p: "South"    for p in ["Nakhon Si Thammarat","Krabi","Phangnga","Phuket","Surat Thani","Ranong","Chumphon","Songkhla","Satun","Trang","Phatthalung","Pattani","Yala","Narathiwat"]},
        }).fillna("Central"))

    # ── Maharashtra: spatial join with OSM district boundaries ────────────────
    mh_df = pd.DataFrame()
    mh_mask = gdf[country_col].str.contains("Maharashtra", case=False, na=False)
    mh = gdf[mh_mask].copy()

    dist_path = Path(__file__).parent / "enrichment_data" / "admin" / "maharashtra_districts.geojson"
    if len(mh) and dist_path.exists():
        try:
            districts = gpd.read_file(str(dist_path)).to_crs("EPSG:4326")
            # Clean district names
            districts["district_name"] = (
                districts["district_name"]
                .str.replace(r"\s+District$", "", regex=True)
                .str.replace(r"\s+Taluka$", "", regex=True)
                .str.strip()
            )
            # Use centroids for the join (faster, avoids geometry complexity)
            mh_pts = mh.copy()
            mh_pts["geometry"] = mh_pts.geometry.centroid
            rc_col_mh = "road_class_norm" if "road_class_norm" in mh_pts.columns else "road_class"
            join_cols = [c for c in ["segment_id", "sss", band_col, rc_col_mh] if c in mh_pts.columns]
            join_cols.append("geometry")
            joined = gpd.sjoin(mh_pts[join_cols],
                               districts[["district_name","geometry"]],
                               how="left", predicate="within")
            joined["district_name"] = joined["district_name"].fillna("Unassigned")
            mh_df = _summarise(joined, "district_name", "Maharashtra PWD / District Authority")
            mh_df = mh_df.rename(columns={"district_name": "District"})
        except Exception as e:
            print(f"  District spatial join skipped: {e}")

    return th_df, mh_df


def export_policy_brief(
    gdf: gpd.GeoDataFrame,
    corridors: gpd.GeoDataFrame,
    output_dir: str,
    top_n: int = 20,   # kept for signature compatibility; no longer the cap
) -> None:
    """
    Export a ministry-ready Excel workbook covering ALL Critical and High Risk segments.

    Sheet 1 — Executive Summary: headline figures by country and band
    Sheet 2 — Critical Segments:  all SSS Critical rows, sorted by priority_index
    Sheet 3 — High Risk Segments: all SSS High Risk rows, sorted by priority_index
    Sheet 4 — Summary by Road Class: counts and avg scores per country x road class
    Sheet 5 — Intervention Zones: corridor-level aggregates (if available)
    Sheet 6 — Methodology Note
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        out_path = Path(output_dir) / "Top_Priority_Interventions.xlsx"

        # ── Filter to scored segments ─────────────────────────────────────────
        mask = (gdf.get("scoreable", pd.Series(False, index=gdf.index)) |
                gdf.get("alignment_scoreable", pd.Series(False, index=gdf.index)))
        scored = gdf[mask].copy()

        sort_col = "priority_index" if "priority_index" in scored.columns else "sss"
        if sort_col in scored.columns:
            scored = scored.sort_values(sort_col, ascending=False)

        # ── Build segment rows for Critical and High Risk ─────────────────────
        band_col = "sss_band" if "sss_band" in scored.columns else "priority_band"
        critical_df_raw = scored[scored[band_col] == "Critical"]
        highrisk_df_raw = scored[scored[band_col] == "High Risk"]

        def build_rows(df_raw, rank_start=1):
            rows = []
            for i, (_, r) in enumerate(df_raw.iterrows()):
                rows.append(_build_segment_row(r, rank_start + i))
            return pd.DataFrame(rows)

        crit_df = build_rows(critical_df_raw, rank_start=1)
        high_df = build_rows(highrisk_df_raw, rank_start=len(crit_df) + 1)

        # ── Sheets 5–6: Admin summaries (needed for summary row counts) ───────
        th_admin_df, mh_admin_df = pd.DataFrame(), pd.DataFrame()
        try:
            th_admin_df, mh_admin_df = _build_admin_summary(scored)
        except Exception as _e:
            print(f"  Admin summary skipped: {type(_e).__name__}: {_e}")

        # ── Sheet 1: Executive Summary ────────────────────────────────────────
        total  = len(scored)
        n_crit = len(critical_df_raw)
        n_high = len(highrisk_df_raw)
        n_mod  = (scored[band_col] == "Moderate").sum()
        n_acc  = (scored[band_col] == "Acceptable").sum()

        country_col = "country" if "country" in scored.columns else "country_code"
        countries   = scored[country_col].unique()

        summary_rows = []
        summary_rows.append(("ADB AI for Safer Roads 2026 — Speed Safety Score", "", "", "", ""))
        summary_rows.append(("Priority Intervention Workbook", "", "", "", ""))
        summary_rows.append(("", "", "", "", ""))
        summary_rows.append(("HEADLINE FIGURES", "", "", "", ""))
        summary_rows.append(("Total scored segments (Tier 2)", total, "", "", ""))
        summary_rows.append(("Critical (immediate action required)", n_crit,
                              f"{100*n_crit/total:.1f}%", "", "SSS >= 60"))
        summary_rows.append(("High Risk (plan within 12 months)", n_high,
                              f"{100*n_high/total:.1f}%", "", "SSS 45–59"))
        summary_rows.append(("Moderate (schedule review)", n_mod,
                              f"{100*n_mod/total:.1f}%", "", "SSS 25–44"))
        summary_rows.append(("Acceptable (no immediate action)", n_acc,
                              f"{100*n_acc/total:.1f}%", "", "SSS < 25"))
        summary_rows.append(("", "", "", "", ""))
        summary_rows.append(("BY COUNTRY", "", "", "", ""))
        for ctry in countries:
            ctry_mask = scored[country_col] == ctry
            ctry_crit = ((scored[country_col] == ctry) & (scored[band_col] == "Critical")).sum()
            ctry_high = ((scored[country_col] == ctry) & (scored[band_col] == "High Risk")).sum()
            avg_sss   = scored.loc[ctry_mask, "sss"].mean() if "sss" in scored.columns else np.nan
            summary_rows.append((ctry, f"{ctry_mask.sum()} segments",
                                  f"{ctry_crit} Critical",
                                  f"{ctry_high} High Risk",
                                  f"Avg SSS {avg_sss:.1f}" if pd.notna(avg_sss) else ""))
        summary_rows.append(("", "", "", "", ""))
        summary_rows.append(("BY ROAD CLASS (Critical + High Risk only)", "", "", "", ""))
        rc_col = "road_class_norm" if "road_class_norm" in scored.columns else "road_class"
        priority_scored = scored[scored[band_col].isin(["Critical", "High Risk"])]
        for rc, grp in priority_scored.groupby(rc_col):
            jur_example = _infer_jurisdiction(rc, str(grp[country_col].mode()[0]) if len(grp) else "")
            summary_rows.append((rc.title(), f"{len(grp)} segments",
                                  f"{(grp[band_col]=='Critical').sum()} Critical",
                                  f"{(grp[band_col]=='High Risk').sum()} High Risk",
                                  jur_example))
        summary_rows.append(("", "", "", "", ""))
        summary_rows.append(("SHEETS IN THIS WORKBOOK", "", "", "", ""))
        summary_rows.append(("Critical Segments",         f"{len(crit_df)} rows", "Immediate action required", "", "Sorted by Priority Index"))
        summary_rows.append(("High Risk Segments",        f"{len(high_df)} rows", "Plan within 12 months",     "", "Sorted by Priority Index"))
        summary_rows.append(("Summary by Road Class",     "Aggregated",           "Country x road class view", "", ""))
        summary_rows.append(("Thailand — By Province",   f"{len(th_admin_df)} provinces", "Risk by province", "", "DOPA province codes"))
        summary_rows.append(("Maharashtra — By District",f"{len(mh_admin_df)} districts", "Risk by OSM district", "", "Spatial join admin_level=5"))
        summary_rows.append(("Methodology Note",          "Reference",            "Column definitions",        "", ""))
        summary_rows.append(("", "", "", "", ""))
        summary_rows.append(("NOTE: All figures are estimates based on ADB-provided sample data. "
                              "Validate against official crash records before policy action.", "", "", "", ""))

        summary_df = pd.DataFrame(summary_rows,
                                  columns=["Item", "Value", "Detail", "Detail 2", "Note"])

        # ── Sheet 4: Summary by Road Class ────────────────────────────────────
        rc_rows = []
        for (ctry, rc), grp in scored.groupby([country_col, rc_col]):
            grp_p   = grp[grp[band_col].isin(["Critical", "High Risk"])]
            avg_sss = grp["sss"].mean() if "sss" in grp.columns else np.nan
            avg_pi  = grp["priority_index"].mean() if "priority_index" in grp.columns else np.nan
            avg_lim = grp["speed_limit"].mean() if "speed_limit" in grp.columns else np.nan
            avg_rec = grp["recommended_limit"].mean() if "recommended_limit" in grp.columns else np.nan
            rc_rows.append({
                "Country":               ctry,
                "Road Class":            rc,
                "Responsible Authority": _infer_jurisdiction(rc, ctry),
                "Total Segments":        len(grp),
                "Critical":              (grp[band_col] == "Critical").sum(),
                "High Risk":             (grp[band_col] == "High Risk").sum(),
                "Moderate":              (grp[band_col] == "Moderate").sum(),
                "Acceptable":            (grp[band_col] == "Acceptable").sum(),
                "Priority Segments":     len(grp_p),
                "% Priority":            f"{100*len(grp_p)/len(grp):.1f}%" if len(grp) else "",
                "Avg SSS":               round(avg_sss, 1) if pd.notna(avg_sss) else "",
                "Avg Priority Index":    round(avg_pi,  1) if pd.notna(avg_pi)  else "",
                "Avg Posted Limit":      round(avg_lim, 0) if pd.notna(avg_lim) else "",
                "Avg Recommended Limit": round(avg_rec, 0) if pd.notna(avg_rec) else "",
                "Avg Reduction Needed":  (round(avg_lim - avg_rec, 0)
                                          if pd.notna(avg_lim) and pd.notna(avg_rec) else ""),
            })
        rc_summary_df = pd.DataFrame(rc_rows)

        # ── Sheet 7: Corridors ────────────────────────────────────────────────
        corr_df = None
        if corridors is not None and len(corridors) > 0:
            keep = [c for c in ["priority_rank", "corridor_label", "country_code",
                                 "n_segments", "sss", "nilsson_fatal_ratio",
                                 "est_lives_saved", "change_effort"]
                    if c in corridors.columns]
            corr_df = corridors[keep].copy()

        # ── Sheet 6: Methodology note ─────────────────────────────────────────
        method_rows = [
            ("Speed Safety Score (SSS)",
             "0–100. Higher = more dangerous misalignment. "
             "Combines: Safe System Alignment (38%), Limit Credibility Gap (30%), VRU Context Risk (32%)."),
            ("SSS Band",
             "Critical ≥60 | High Risk 45–59 | Moderate 25–44 | Acceptable <25"),
            ("Priority Band",
             "Exposure × Likelihood × Severity index. Determines order within a band. "
             "Critical ≥57 | High Risk 47–57 | Moderate 37–47 | Acceptable <37."),
            ("Responsible Authority",
             "Inferred from road class and country. Motorway/trunk = national highway authority. "
             "Primary = state/provincial roads. Secondary = rural roads dept. Residential = local government. "
             "Verify with official gazetteer before correspondence."),
            ("Recommended Limit (km/h)",
             "WHO Safe System speed tier for this road's class and land use, "
             "reduced for sinuous geometry (AASHTO Green Book method)."),
            ("Reduction Needed (km/h)",
             "Posted limit minus recommended limit. Positive = posted limit should be reduced."),
            ("Intervention Actions",
             "Specific engineering actions derived from score drivers. "
             "Speed limit change (SSS alignment gap); curve treatment (sinuosity ≥1.20); "
             "enforcement camera (Mapillary blindspot); lighting (OSM lit=no in urban area or high VIIRS NTL); "
             "resurfacing (OSM surface unpaved); median barrier (Nilsson ratio >4 on trunk/primary); "
             "traffic calming (residential road >30 km/h)."),
            ("Nilsson Fatal Risk Ratio",
             "Ratio of estimated fatal crash risk at observed speed vs. Safe System speed. "
             "Based on Nilsson Power Model (WHO endorsed): risk ∝ (speed/reference)^4.5."),
            ("Sinuosity Index",
             "Path length / straight-line distance. 1.0 = straight road. "
             "≥1.20 triggers advisory sign recommendation. ≥1.50 triggers chevron warning signs."),
            ("VIIRS NTL Score",
             "Normalized VIIRS 2025 nighttime radiance (0–100). "
             "Values >60 indicate significant after-dark pedestrian or commercial activity near the road."),
            ("Mapillary Blindspot",
             "Segment flagged as high-risk (SSS ≥45) with no Mapillary street imagery coverage — "
             "no remote monitoring or enforcement baseline available."),
            ("OSM Surface / OSM Lit",
             "Physical road attributes from OpenStreetMap road infrastructure extract (Geofabrik 2025 PBF). "
             "'unknown' = not recorded in OSM for this segment."),
            ("Street View Link",
             "Google Maps Street View coordinates for the segment midpoint. "
             "Open in browser to visually verify conditions before engineering review."),
            ("Data source",
             "ADB AI for Safer Roads 2026 challenge dataset. GPS speed data from ADB-provided "
             "Maharashtra and Thailand GeoJSON files. OSM, VIIRS, WorldPop, HOTOSM are open-source global datasets."),
            ("Validation",
             "No crash outcome validation — no crash location data was provided. "
             "Scores are calibrated against the WHO Safe System framework and validated via "
             "5-fold XGBoost CV (RMSE=7.95, R²=0.817) and weight sensitivity testing "
             "(ρ>0.95 across 600 weight perturbations). "
             "Validate against official crash records before policy action."),
        ]
        method_df = pd.DataFrame(method_rows, columns=["Column / Term", "Explanation"])

        # ── Write workbook ────────────────────────────────────────────────────
        FILL_CRIT   = PatternFill("solid", fgColor="FFCCCC")
        FILL_HIGH   = PatternFill("solid", fgColor="FFE5CC")
        FILL_HDR    = PatternFill("solid", fgColor="002569")
        FONT_HDR    = Font(bold=True, color="FFFFFF", size=10)
        THIN_SIDE   = Side(style="thin", color="CCCCCC")

        def style_header_row(ws, n_cols):
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = FILL_HDR
                cell.font = FONT_HDR
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        def autofit(ws, df_s, max_w=55):
            for col_idx, col in enumerate(df_s.columns, start=1):
                vals = df_s[col].astype(str).str.len()
                best = max(len(str(col)), int(vals.max()) if len(vals) else 0)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(best + 2, max_w)

        def freeze_and_filter(ws):
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)

            if len(crit_df):
                crit_df.to_excel(writer, sheet_name="Critical Segments", index=False)
                ws = writer.sheets["Critical Segments"]
                style_header_row(ws, len(crit_df.columns))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = FILL_CRIT
                autofit(ws, crit_df)
                freeze_and_filter(ws)

            if len(high_df):
                high_df.to_excel(writer, sheet_name="High Risk Segments", index=False)
                ws = writer.sheets["High Risk Segments"]
                style_header_row(ws, len(high_df.columns))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.fill = FILL_HIGH
                autofit(ws, high_df)
                freeze_and_filter(ws)

            rc_summary_df.to_excel(writer, sheet_name="Summary by Road Class", index=False)
            ws = writer.sheets["Summary by Road Class"]
            style_header_row(ws, len(rc_summary_df.columns))
            autofit(ws, rc_summary_df)

            if len(th_admin_df):
                th_admin_df.to_excel(writer, sheet_name="Thailand — By Province", index=False)
                ws = writer.sheets["Thailand — By Province"]
                style_header_row(ws, len(th_admin_df.columns))
                autofit(ws, th_admin_df)
                freeze_and_filter(ws)

            if len(mh_admin_df):
                mh_admin_df.to_excel(writer, sheet_name="Maharashtra — By District", index=False)
                ws = writer.sheets["Maharashtra — By District"]
                style_header_row(ws, len(mh_admin_df.columns))
                autofit(ws, mh_admin_df)
                freeze_and_filter(ws)

            if corr_df is not None:
                corr_df.to_excel(writer, sheet_name="Intervention Zones", index=False)

            method_df.to_excel(writer, sheet_name="Methodology Note", index=False)
            ws = writer.sheets["Methodology Note"]
            style_header_row(ws, 2)
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 90
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[1].height = 18

        print(f"\n  Policy brief exported: {out_path.name}")
        print(f"  Critical: {len(crit_df)} segments  |  High Risk: {len(high_df)} segments")
        print(f"  Total priority interventions: {len(crit_df) + len(high_df)}")

    except ImportError:
        print("  Policy brief skipped — pip install openpyxl to enable")
    except Exception as e:
        import traceback
        print(f"  Policy brief export failed — {e}")
        traceback.print_exc()
